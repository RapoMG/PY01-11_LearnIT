"""Homework for lesson 9"""

from functools import wraps
from re import split as respl  # task 2
import json  # task 3, 4 and 10
import csv  # task 5 and 6
from pathlib import Path  # task 7 and 10
from openpyxl import Workbook, load_workbook  # task 9 
import os  # task 10

def task_separator(last=False):
    """
    Separates successive functions called after each other with header and input() based pause at the end of its execution.

    :param last: if True omits pause at the end. *Default:* False 
    """

    def wrapper(func):

        @wraps(func)
        def call(*args, **kwargs):
            line_l = '<<< '
            line_r = ' >>>'
            title = func.__name__
            instr = func.__doc__

            title = title.replace("task", "task ")
            
            print(f"\n{line_l * 2} {title.upper()} {line_r * 2}\n")
            print(f"Instructions: \n \t{instr}\n\nCode execution:\n")

            result = func(*args, **kwargs)

            if not last: input("\nPress [Enter] for the next function\n") 

            return result
        
        return call

    return wrapper


########## TASKS #########

@task_separator()
def task1():
    """Dziennik użytkownika: Napisz program, który w pętli prosi użytkownika o wpisanie jednej
linii tekstu. Każda wpisana linia powinna być dopisywana (tryb 'a' ) do pliku
dziennik.txt . Program kończy działanie, gdy użytkownik wpisze "koniec"."""

    while True:
        logging = input("Enter text to log it or [end] to end:\n")
        if logging.lower().strip() == "end":
            print("\nLogging stoped.")
            break
        with open("dziennik.txt", "a", encoding="utf-8") as f:
            f.write(f"{logging}\n")


@task_separator()
def task2():
    """Licznik słów: Stwórz program, który pyta o nazwę pliku, odczytuje go, a następnie zlicza i
wyświetla całkowitą liczbę słów w tym pliku. Obsłuż błąd FileNotFoundError , jeśli plik nie
istnieje."""

    #from re import split as respl

    while True:
        words = []
        file = input("Enter file name [example.txt] to count words or [end] to exit:\n")
        if file.strip().lower() == "end":
            break
        try:
            with open(file, "r", encoding="utf-8") as f:
                # sprit on multiple delimiters, but creates empty strings
                words = respl(r"[,;:.\s\\()\?\!]", f.read())
        except FileNotFoundError:
            print(f"\nNo file named {file}\n")
            continue

        #counting minus empty elements
        count = len(words) - words.count("")

        s = "" if count == 1 else "s"
        print(f"\nThere are {count} word{s} in the {file} file.\n")

   
@task_separator()
def task3():
    """Konfiguracja w JSON: Stwórz słownik Pythona z ustawieniami aplikacji, np.
konfiguracja = {"uzytkownik": "admin", "motyw": "ciemny", "rozdzielczosc":
[1920, 1080]} . Zapisz ten słownik do pliku config.json z wcięciami i poprawnym
kodowaniem polskich znaków."""

    #import json
    
    config  = {
        "user_rank": "admin",
        "user_name": "Zośka",
        "theme": "dark",
        "resolution":[1920, 1080],
    }
    
    with open("config.json", "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4, ensure_ascii=False)
        print("Configuration saved.")


@task_separator()
def task4():
    """Odczyt konfiguracji: Napisz program, który odczytuje plik config.json z poprzedniego
zadania i wyświetla komunikat: Witaj, [uzytkownik]! Twój motyw to [motyw]."""

    #import json
    try:
        with open("config.json", "r", encoding="utf-8") as f:
            config = json.load(f)
    except FileNotFoundError:
       print("No configuration file!")
    
    print(f"Hello, {config.get("user_name")}! Theme set: {config.get("theme")}.")
    
@task_separator()
def task5():
    """Eksport do CSV: Masz listę słowników: produkty = [{"nazwa": "Mleko", "cena":
3.50}, {"nazwa": "Chleb", "cena": 4.20}] . Zapisz te dane do pliku produkty.csv ,
gdzie pierwszy wiersz to nagłówki ("nazwa", "cena")."""

    #import csv

    items = [
        {
            "name": "milk",
            "price": 3.5,
        },
        {
            "name": "bread",
            "price": 8,
        },

    ]
    header = ["name", "price"]

    with open("items.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        exp_items = [header]
        for item in items:
            exp_items.append([item.get("name"), item.get("price")]) 
        
        writer.writerows(exp_items)
        print("Items list exported.")
    

@task_separator()
def task6():
    """Import z CSV: Napisz program, który odczytuje plik produkty.csv i oblicza sumę cen
wszystkich produktów. Użyj csv.DictReader , aby łatwiej odwoływać się do kolumn po
nazwach."""

    #import csv

    with open("items.csv", "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=",")
        imp_items = list(reader)
    
    total = None
    for item in imp_items:
        if total is None:
            total = 0
        total += float(item.get("price"))
    
    print(f"Total price for imported items: {total}")
   

@task_separator()
def task7():
    """Tworzenie struktury folderów: Użyj modułu pathlib , aby napisać skrypt, który tworzy
strukturę folderów: Projekt/src , Projekt/data , Projekt/docs ."""
    
    #from pathlib import Path

    main = Path("Project")
    folders = ["src", "data", "docs"]

    for folder in folders:
        path = main / folder

        # chceck if whole path exists
        if not path.is_dir():
            try:
                # parameter parents=True creates parent folder if it doesn't exist
                # parameter exist_ok=True ignores error if already exists
                path.mkdir(parents=True, exist_ok=True)
            except PermissionError:
                print("Unable to create folder: permission denied.")
            except Exception as e:
                print(f"An error occurred: {e}")
            
            print(f"Directory {path} created.")
        else:
            print(f"Directory {path} already exist.")



@task_separator()
def task8():
    """Wyszukiwarka logów: Wyobraź sobie, że masz duży plik log.txt . Napisz program, który
pyta użytkownika o szukane słowo (np. "ERROR") i zapisuje wszystkie linie zawierające to
słowo do nowego pliku wyniki_wyszukiwania.txt ."""
    
    # Definitions
    errors = []
    count = 0
    exit_log = ""
    logs = []


    try:
        with open("log_with_errors_L9.txt", "r", encoding="utf-8") as f:
            logs = f.readlines()
    except FileNotFoundError:
        print("No log file or incorrect file name.")

    
    for line in logs:
        if "ERROR" in line:
            count += 1
            errors.append(line)
    if count > 0:
        with open("log_only_errors_L9", "w", encoding="utf-8") as f:
            f.writelines(errors)
            exit_log = "File with all errors was created."

    s = "" if count == 1 else "s"
    if len(logs) > 0:
        print(f"{count} error{s} found in the log. {exit_log}")


@task_separator()
def task9():
    """Prosty arkusz kalkulacyjny: Używając openpyxl , stwórz plik finanse.xlsx . W
pierwszej kolumnie umieść nazwy wydatków (np. "Czynsz", "Jedzenie"), a w drugiej ich
wartości. W komórce poniżej wartości oblicz i wstaw sumę wszystkich wydatków, używając
formuły Excela (np. =SUM(B1:B2) )."""

    #from openpyxl import Workbook, load_workbook

    file = "finances_L9.xlsx"

    wb = Workbook()
    ws = wb.active  # active sheet
    ws.title = "Expenses"

    rows = ["Rent", "Food", "Bills", "Hobby"]
    values = [810, 1200, 400, 221]

    for row, value in zip(rows, values):
        ws.append([row, value])

    ws["B5"] = "=SUM(B1:B4)"

    wb.save(file)

    print(f"File {file} created.")


@task_separator(True)
def task10():
    """Mini-projekt: Lista zadań: Stwórz prostą aplikację do zarządzania listą zadań. Program
powinien:
Przy starcie próbować wczytać zadania z pliku zadania.json .
Pozwalać użytkownikowi dodać nowe zadanie.
Pozwalać wyświetlić wszystkie zadania.
Przy zamknięciu (lub na polecenie) zapisywać aktualną listę zadań do pliku
zadania.json ."""
    
    # Would be better with datetime
    # import json
    # from pathlib import Path
    # import os

    # Pretend to be config
    def file_name():
        return "zadania.json"
    
    def choices():
        options_list = {
            "1": "    1.  Show tasks",
            "2": "    2.  Add new task",
            "3": "    3.  Save data",
            "0": "    0.  Exit app",
        }

        return options_list
    
    def clean():  # Clean screen
        """Removes all characters from terminal window."""
        print(f"{'*'*15}  Clean screen doesn't work here {'*'*15}")
        os.system('cls' if os.name == 'nt' else 'clear')
        print('')


    class Task:
        """
        Represents a single task with a name, date, optional description,
        and completion status.
        """

        def __init__(self, task_name: str, task_date: str, description="", finished=False):
            """
            Initialize a new Task instance.

            :param task_name: Human-readable name of the task.
            :param task_date: Date associated with the task (e.g., due date),represented as a string.
            :param description: Optional detailed description of the task.
            :param finished: Mark the task as completed. Should be ``False`` for the new task.
            """
            self.task_name = task_name
            self.task_date = task_date
            self.description = description
            self.finished = finished

        def json(self) -> dict:
            """
            Serialize the task into a dictionary representation.

            :return: A dictionary containing task attributes suitable for JSON serialization.
            """
            return {
                "task_name": self.task_name,
                "task_date": self.task_date,
                "description": self.description,
                "finished": self.finished,
            }
        
        def mark_finished(self):
            """
            Mark the task as completed.
            """
            self.finished = True

        def __repr__(self):
            """
            Return an unambiguous string representation of the Task instance,
            primarily for debugging purposes.

            :return: String representation of the task.
            """
            return f"Task('{self.task_name}', '{self.task_date}','{self.description}','{self.finished}')"


    class Organiser:
        """
        Represents a collection of tasks grouped under a common name.
        """
        def __init__(self, group_name: str):
            """
            Initialize a new Organiser instance.

            :param group_name: Name of the task group (e.g., project or category).
            """
            self.group = group_name
            self.all_tasks = []

        def new_task(self, task_name: str, task_date: str, description= "", finished=False):
            """
            Create a new task and add it to the organiser.

            :param task_name: Name of the task.
            :param task_date: Date associated with the task.
            :param description: Optional task description.
            :param  finished: Marker for finished tasks
            """
            self.all_tasks.append(Task(task_name, task_date, description, finished))
        
        def json(self):
            """
            Serialize the organiser and its tasks into a dictionary representation.

            :return: A dictionary containing the group name and a list of serialized tasks.
            """
            return {
                'group': self.group,
                'tasks': [task.json() for task in self.all_tasks]
            }
        
        def __repr__(self):
            """
            Return a concise string representation of the Organiser instance.

            :return: String representation showing the group name and task count.
            """
            return f'Organiser({self.group}, tasks: {len(self.all_tasks)})'


    def load_file() -> dict | None:
        """
        Load and deserialize JSON data from a file.

        The function determines the target file path using ``file_name()``.
        If the file exists, its contents are read and parsed as JSON.
        If the file does not exist, the function returns ``None``.

        :return: A dictionary containing the deserialized JSON data if the
                file exists; otherwise, ``None``.
        """

        file = Path(file_name())

        # chceck if file already exists
        if file.exists():
            # load file
            try:
                with open(file_name(), "r", encoding="utf-8") as f:
                    return json.load(f)        
            except json.decoder.JSONDecodeError:
                return None
        else:            
           return None 
        

    def save_file(data: dict):
        """
        Serialize data to JSON and save it to a file.

        The function determines the target file path using ``file_name()``
        and writes the provided data object to the file in JSON format.
        Existing content will be overwritten.

        :param data: The data dictionary to serialize. It must be JSON-serializable.
        """
        with open(file_name(), "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)


    def restorer(dicts: dict) -> object:
        """
        Docstring for restorer
        
        :param dicts: Description
        :type dicts: dict
        :return: Description
        :rtype: object
        """

        # create Organiser class instance
        restored = Organiser(dicts.get("group"))
        # get tasks list
        tasks = dicts.get("tasks")
        # for each task create task instance
        for task in tasks:
            # task is dict
            # call each element and send it to new task
            restored.new_task(
                task.get("task_name"),
                task.get("task_date"), 
                task.get("description"),
                task.get("finished")
                )
        # return Organiser class instance
        return restored
    
    def main_screen(tasks_number) -> int:
        options = choices()
        while True:
            # clean screen
            clean()
            # inform about number of tasks
            if tasks_number == 0:
                options.pop("1")
                print("\n    You have no tasks in this organiser.\n")
            else:
                s = "" if tasks_number == 1 else "s"
                print(f"\n    You have {tasks_number} task{s} in this organiser.\n")

            # print optios 
            for option in options.keys():
                print(options.get(option))
        
            # ask to chose
            choice = input("\n    What you want to do?\n")
            if choice.strip() in options.keys():
                return int(choice)
            
        
    def show_tasks(tasks: list):
        clean()
        # task is a list of class Task objects
        counter = 1
        for task in tasks:
            print(f"    {counter}. {task.task_name}  due to: {task.task_date}  Finished: {task.finished}")
            print(f"         Description:\n        {task.description}\n")
            counter += 1
            
        input("Press [ENTER] to continue\n1")

    def add_task(organizer: object):
        clean()
        name = input("\n    Name your task:\n")
        date  = input("    Deadline for this task:\n")
        description  = input("    Additional information:\n")

        organizer.new_task(name, date, description)

    
    def app():
        
        # call loader
        loaded_data = load_file()
        if loaded_data is None:
            # create main class
            main_org = Organiser("Main")
        else:
            # restore data
            main_org = restorer(loaded_data)

        # place for a loop
        while True:
            # Print options
            tasks_num = len(main_org.all_tasks) # number of tasks
            option = main_screen(tasks_num)
            
            # show all tasks
            if option == 1:
                show_tasks(main_org.all_tasks)
            # add new task
            elif option == 2:
                add_task(main_org)
            #     save tasks
            elif option == 3:
                save_file(main_org.json())
            #     exit app
            elif option == 0:
                # auto save
                save_file(main_org.json())
                break


    input("Fake pause before running app(). Press [ENTER] to continue")
    app()    

if __name__ == "__main__":
    task1()
    task2()
    task3()
    task4()
    task5()
    task6()
    task7()
    task8()
    task9()
    task10()
 